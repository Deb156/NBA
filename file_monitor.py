import os
import json
import time
import shutil
import smtplib
import ssl
import threading
from datetime import datetime, timedelta
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.image import MIMEImage
# Removed watchdog imports due to Python 3.13 compatibility issues
from jinja2 import Template
from openpyxl import Workbook, load_workbook
from email.mime.base import MIMEBase
from email import encoders



class FileMonitor:
    def __init__(self, config_path='config.json'):
        with open(config_path, 'r') as f:
            self.config = json.load(f)
        
        self.dropped_files = []
        self.transfer_log = []
        self.performance_tracking = {}  # Track each asset's transfer progress
        self.performance_data = []  # Store performance analysis data
        self.running = True
        self.excel_file = 'NBA_Transfer_Performance.xlsx'
        self.serial_counter = 1
        self.transfer_status_cache = {}  # Cache transfer status to detect changes
        self.last_status_alert_time = datetime.now()  # Track last alert time
        self.file_duplication_tracker = {}  # Track file duplications and overwrites
        self.recent_alerts = {}  # Track recent alerts to prevent duplicates
        
        # Create directories if they don't exist
        for folder in self.config['watch_folders'].values():
            os.makedirs(folder, exist_ok=True)
        for folder in self.config['destination_folders'].values():
            os.makedirs(folder, exist_ok=True)

    def count_files_in_folder(self, folder_path):
        if not os.path.exists(folder_path):
            return 0
        count = 0
        for root, dirs, files in os.walk(folder_path):
            count += len(files)
        return count

    def is_transfer_ongoing(self, file_path):
        # Check if file size is changing (simple transfer detection)
        try:
            size1 = os.path.getsize(file_path)
            time.sleep(0.1)
            size2 = os.path.getsize(file_path)
            return size1 != size2
        except:
            return False

    # The script will print email alerts to console instead of sending emails when the password is set to "your_password"
    # if sender_password="your_password" in config.json

    def get_asset_type(self, filename):
        ext = os.path.splitext(filename)[1].lower()
        if ext in ['.mp4', '.avi', '.mov', '.mkv', '.wmv']:
            return 'Video'
        elif ext in ['.jpg', '.jpeg', '.png', '.gif', '.bmp']:
            return 'Image'
        elif ext in ['.mp3', '.wav', '.flac', '.aac']:
            return 'Audio'
        elif ext in ['.txt', '.doc', '.docx', '.pdf', '.xml']:
            return 'Document'
        else:
            return 'File'
    
    def track_asset_drop(self, asset_name, source_folder, drop_time, is_folder=False):
        asset_key = f"{source_folder}_{asset_name}"
        asset_type = 'Folder' if is_folder else self.get_asset_type(asset_name)
        
        self.performance_tracking[asset_key] = {
            'sl_no': self.serial_counter,
            'asset_name': asset_name,
            'asset_type': asset_type,
            'source_location': source_folder,
            'upload_time': drop_time,
            'destinations': {},
            'remarks': ''
        }
        self.serial_counter += 1
    
    def update_transfer_status(self, asset_name, source_folder, dest_name, transfer_time=None):
        asset_key = f"{source_folder}_{asset_name}"
        if asset_key in self.performance_tracking:
            if transfer_time:
                self.performance_tracking[asset_key]['destinations'][dest_name] = transfer_time
            else:
                self.performance_tracking[asset_key]['destinations'][dest_name] = 'In Progress'
    
    def update_performance_data(self, asset_name, asset_type, count, start_time, transfer_times):
        end_time = max(transfer_times.values()) if transfer_times else start_time
        total_time = (end_time - start_time).total_seconds() / 60  # in minutes
        
        self.performance_data.append({
            'Asset Name': asset_name,
            'Asset Type': asset_type,
            'Count': count,
            'Start Time': start_time.strftime('%Y-%m-%d %H:%M:%S'),
            'End Time': end_time.strftime('%Y-%m-%d %H:%M:%S'),
            'Total Time (minutes)': round(total_time, 2),
            'Transfer Details': ', '.join([f"{k}: {v.strftime('%H:%M:%S')}" for k, v in transfer_times.items()])
        })
    
    def create_excel_report(self):
        try:
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            wb = Workbook()
            ws = wb.active
            ws.title = 'NBA Transfer Performance'
            
            # Headers with formatting
            headers = ['SL.NO', 'Asset Name', 'Asset Type', 'Source Drop Location', 
                      'Asset Upload Time', 'Destination Location', 'Asset Transferred Time', 
                      'Time Difference Analysis', 'Remarks']
            ws.append(headers)
            
            # Format headers - bold, background color, alignment
            header_font = Font(bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center')
            
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Set column widths
            column_widths = [8, 25, 12, 20, 20, 18, 20, 20, 30]
            for i, width in enumerate(column_widths, 1):
                ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width
            
            # Add data rows
            for asset_key, data in self.performance_tracking.items():
                destinations = self.config['destination_mapping'][data['source_location']]
                
                for dest_name in destinations:
                    transfer_time = data['destinations'].get(dest_name, 'In Progress')
                    
                    if transfer_time != 'In Progress' and hasattr(transfer_time, 'strftime'):
                        time_diff = (transfer_time - data['upload_time']).total_seconds() / 60
                        time_analysis = f"{time_diff:.2f} minutes"
                        transfer_time_str = transfer_time.strftime('%Y-%m-%d %H:%M:%S')
                    else:
                        time_analysis = 'In Progress'
                        transfer_time_str = 'In Progress'
                    
                    ws.append([
                        data['sl_no'],
                        data['asset_name'],
                        data['asset_type'],
                        data['source_location'],
                        data['upload_time'].strftime('%Y-%m-%d %H:%M:%S'),
                        dest_name,
                        transfer_time_str,
                        time_analysis,
                        data['remarks']
                    ])
            
            # Add borders to all cells with data
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(headers)):
                for cell in row:
                    cell.border = thin_border
                    if cell.row > 1:  # Data rows alignment
                        cell.alignment = Alignment(horizontal='left', vertical='center')
            
            wb.save(self.excel_file)
        except Exception as e:
            print(f"Excel creation failed: {e}")
    
    def send_email(self, subject, body, alert_type="Alert", details=None, attach_excel=False):
        try:
            # Rate limiting: prevent duplicate alerts within 30 seconds
            alert_key = f"{subject}_{body[:50]}"
            current_time = datetime.now()
            if alert_key in self.recent_alerts:
                time_diff = (current_time - self.recent_alerts[alert_key]).total_seconds()
                if time_diff < 30:
                    return
            self.recent_alerts[alert_key] = current_time
            
            if self.config['email']['sender_password'] == 'your_password':
                print(f"EMAIL ALERT: {subject}\n{body}\n")
                return
            
            # Select appropriate template based on alert type
            if "Transfer Monitoring Notification" in subject:
                template_file = 'melts_transfer_report_template.html'
            elif "Transfer Status" in subject:
                template_file = 'transfer_status_template.html'
            elif "Performance Report" in subject:
                template_file = 'performance_report_template.html'
            elif "Drop Alert Notification" in subject:
                template_file = 'melts_drop_alert_template.html'
            else:
                template_file = 'email_template.html'
            
            # Load and render HTML template
            with open(template_file, 'r', encoding='utf-8') as f:
                template = Template(f.read())
            
            alert_class = "danger" if "Alert" in alert_type else "info" if "Report" in alert_type else "success"
            
            html_content = template.render(
                subject=subject,
                message=body,
                alert_type=alert_type,
                alert_class=alert_class,
                details=details,
                timestamp=datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            )
            
            # Use 'related' only when attaching logo, otherwise use 'alternative'
            if "Drop Alert Notification" in subject or "Transfer Monitoring Report" in subject or "Transfer Status" in subject or "Performance Report" in subject:
                msg = MIMEMultipart('related')
                msg['From'] = self.config['email']['sender_email']
                msg['To'] = ', '.join(self.config['email']['recipients'])
                msg['Subject'] = subject
                
                # Create alternative container for text and HTML
                msg_alternative = MIMEMultipart('alternative')
                msg_alternative.attach(MIMEText(body, 'plain', 'utf-8'))
                msg_alternative.attach(MIMEText(html_content, 'html', 'utf-8'))
                msg.attach(msg_alternative)
                
                # Attach appropriate logo based on email type
                try:
                    if "Performance Report" in subject:
                        # Attach CLEAR NBA Logo for performance reports
                        with open('CLEAR NBA Logo.png', 'rb') as f:
                            img_data = f.read()
                        image = MIMEImage(img_data)
                        image.add_header('Content-ID', '<nba_logo>')
                        image.add_header('Content-Disposition', 'inline', filename='nba_logo.png')
                        msg.attach(image)
                    else:
                        # Attach CLEAR logo for other emails
                        with open('CLEAR Logo.png', 'rb') as f:
                            img_data = f.read()
                        image = MIMEImage(img_data)
                        image.add_header('Content-ID', '<clear_logo>')
                        image.add_header('Content-Disposition', 'inline', filename='clear_logo.png')
                        msg.attach(image)
                except Exception as e:
                    print(f"Failed to attach logo: {e}")
            else:
                msg = MIMEMultipart('alternative')
                msg['From'] = self.config['email']['sender_email']
                msg['To'] = ', '.join(self.config['email']['recipients'])
                msg['Subject'] = subject
                
                msg.attach(MIMEText(body, 'plain', 'utf-8'))
                msg.attach(MIMEText(html_content, 'html', 'utf-8'))
            
            # Attach Excel file if requested
            if attach_excel and os.path.exists(self.excel_file):
                with open(self.excel_file, 'rb') as f:
                    part = MIMEBase('application', 'octet-stream')
                    part.set_payload(f.read())
                    encoders.encode_base64(part)
                    part.add_header('Content-Disposition', f'attachment; filename="{self.excel_file}"')
                    msg.attach(part)
            
            context = ssl.create_default_context()
            server = smtplib.SMTP(self.config['email']['smtp_server'], self.config['email']['smtp_port'])
            server.starttls(context=context)
            server.login(self.config['email']['sender_email'], self.config['email']['sender_password'])
            server.send_message(msg)
            server.quit()
        except Exception as e:
            print(f"Email send failed: {e}")
            print(f"EMAIL ALERT: {subject}\n{body}\n")



def polling_worker(monitor):
    folder_snapshots = {}
    processed_files = set()  # Track already processed files
    
    # Initialize snapshots
    for folder_name, folder_path in monitor.config['watch_folders'].items():
        folder_snapshots[folder_name] = set()
        if os.path.exists(folder_path):
            for root, dirs, files in os.walk(folder_path):
                for item in files + dirs:
                    item_path = os.path.join(root, item)
                    folder_snapshots[folder_name].add(item_path)
                    processed_files.add(item_path)  # Mark as already processed
    
    while monitor.running:
        time.sleep(2)  # Check every 2 seconds
        
        for folder_name, folder_path in monitor.config['watch_folders'].items():
            if not os.path.exists(folder_path):
                continue
                
            current_items = set()
            for root, dirs, files in os.walk(folder_path):
                for item in files + dirs:
                    item_path = os.path.join(root, item)
                    current_items.add(item_path)
            
            # Check for new items
            new_items = current_items - folder_snapshots[folder_name]
            
            for file_path in new_items:
                if not os.path.exists(file_path) or file_path in processed_files:
                    continue
                processed_files.add(file_path)  # Mark as processed
                    
                # Record the drop
                file_count = 0
                if os.path.isdir(file_path):
                    file_count = monitor.count_files_in_folder(file_path)
                    if file_count == 0:
                        monitor.send_email(
                            "Blank Folder Alert",
                            f"Blank folder detected:\nSource: {folder_name}\nFolder: {os.path.basename(file_path)}\nTime: {datetime.now()}"
                        )
                
                drop_time = datetime.now()
                asset_name = os.path.basename(file_path)
                
                # Check for file duplication/overwrite
                duplication_key = f"{folder_name}_{asset_name}"
                file_size = 0
                try:
                    if os.path.isfile(file_path):
                        file_size = os.path.getsize(file_path)
                except:
                    file_size = 0
                
                # Track file duplication
                if duplication_key in monitor.file_duplication_tracker:
                    tracker = monitor.file_duplication_tracker[duplication_key]
                    tracker['duplication_count'] += 1
                    tracker['current_size'] = file_size
                    tracker['current_drop_time'] = drop_time
                    
                    file_type = monitor.get_asset_type(asset_name) if os.path.isfile(file_path) else 'Folder'
                    duplicate_version = f"V{tracker['duplication_count'] + 1}"
                    
                    duplication_alert = {
                        'File Name': asset_name,
                        'File Type': file_type,
                        'Size of Previous File (bytes)': tracker['original_size'],
                        'Size of Current File (bytes)': file_size,
                        'Previous File Drop Time': tracker['original_drop_time'].strftime('%Y-%m-%d %H:%M:%S'),
                        'Current File Drop Time': drop_time.strftime('%Y-%m-%d %H:%M:%S'),
                        'Duplicate Version': duplicate_version,
                        'Source Location': folder_name
                    }
                    
                    monitor.send_email(
                        "Transfer Status Alert - Immediate Action Required",
                        f"Duplicate file detected: {asset_name} ({duplicate_version}) dropped in {folder_name} source folder.",
                        "Transfer Status Alert",
                        {'Duplicate File Detection': [duplication_alert]}
                    )
                else:
                    monitor.file_duplication_tracker[duplication_key] = {
                        'original_size': file_size,
                        'original_drop_time': drop_time,
                        'current_size': file_size,
                        'current_drop_time': drop_time,
                        'duplication_count': 0
                    }
                
                monitor.dropped_files.append({
                    'time': drop_time,
                    'source_folder': folder_name,
                    'asset_name': asset_name,
                    'asset_type': 'folder' if os.path.isdir(file_path) else 'file',
                    'file_count': file_count,
                    'path': file_path
                })
                
                monitor.track_asset_drop(asset_name, folder_name, drop_time, os.path.isdir(file_path))
                
                details = {
                    'Source Folder': folder_name,
                    'Asset Name': os.path.basename(file_path),
                    'Asset Type': 'Folder' if os.path.isdir(file_path) else 'File',
                    'File Count': file_count if os.path.isdir(file_path) else 1
                }
                
                monitor.send_email(
                    f"NBA MELTS Drop Alert Notification - {folder_name}",
                    f"New {details['Asset Type'].lower()} detected in {folder_name} watch folder.",
                    "NBA MELTS Drop Alert Notification",
                    details
                )
            
            folder_snapshots[folder_name] = current_items

def validation_worker(monitor):
    while monitor.running:
        time.sleep(monitor.config['intervals']['validation_minutes'] * 60)
        
        current_time = datetime.now()
        status_alerts = []
        immediate_alerts = []
        
        # Check transfer status for all dropped files
        for item in monitor.dropped_files:
            asset_key = f"{item['source_folder']}_{item['asset_name']}"
            source_exists = os.path.exists(item['path'])
            time_since_drop = (current_time - item['time']).total_seconds() / 60  # minutes
            
            destinations = monitor.config['destination_mapping'][item['source_folder']]
            
            for dest_name in destinations:
                dest_path = monitor.config['destination_folders'][dest_name]
                expected_file = os.path.join(dest_path, item['asset_name'])
                dest_exists = os.path.exists(expected_file)
                
                # Determine status and RCA
                status = 'Pending'
                rca = 'Monitoring'
                alert_type = 'scheduled'
                
                if dest_exists:
                    status = 'Transferred'
                    rca = 'Successfully Completed'
                    # Only update transfer time if not already recorded
                    if (asset_key in monitor.performance_tracking and 
                        dest_name not in monitor.performance_tracking[asset_key]['destinations']):
                        transfer_time = datetime.now()
                        monitor.update_transfer_status(item['asset_name'], item['source_folder'], dest_name, transfer_time)
                elif source_exists and time_since_drop > monitor.config['transfer_status_thresholds']['no_transfer_threshold_minutes'] and not dest_exists:
                    status = 'Failed'
                    rca = 'Transfer Failed - File Present in Source but Not Transferred'
                    alert_type = 'immediate'
                elif source_exists and time_since_drop > monitor.config['transfer_status_thresholds']['transfer_progress_threshold_minutes'] and not dest_exists:
                    # Check for intermittent transfer failure
                    if asset_key in monitor.transfer_status_cache and monitor.transfer_status_cache[asset_key].get(dest_name) == 'In Progress':
                        status = 'Failed'
                        rca = 'Intermittent Transfer Failure - Transfer was in progress but failed'
                        alert_type = 'immediate'
                    else:
                        status = 'In Progress'
                        rca = 'Transfer in Progress'
                        alert_type = 'immediate'
                        # Cache the in-progress status
                        if asset_key not in monitor.transfer_status_cache:
                            monitor.transfer_status_cache[asset_key] = {}
                        monitor.transfer_status_cache[asset_key][dest_name] = 'In Progress'
                
                # Create alert entry if conditions are met
                if status in ['Failed', 'In Progress'] and rca != 'Successfully Completed':
                    alert_entry = {
                        'Asset Name': item['asset_name'],
                        'Asset Type': item['asset_type'].title(),
                        'Source Location': item['source_folder'],
                        'Destination Location': dest_name,
                        'Status': status,
                        'Root Cause Analysis': rca,
                        'Duration (minutes)': f"{time_since_drop:.1f}",
                        'Drop Time': item['time'].strftime('%Y-%m-%d %H:%M:%S'),
                        'Log Analysis': f"Source exists: {source_exists}, Destination exists: {dest_exists}, Time elapsed: {time_since_drop:.1f} min"
                    }
                    
                    if alert_type == 'immediate':
                        immediate_alerts.append(alert_entry)
                    else:
                        status_alerts.append(alert_entry)
        
        # Send immediate alerts
        if immediate_alerts:
            monitor.send_email(
                "Transfer Status Alert - Immediate Action Required",
                f"Critical transfer issues detected for {len(immediate_alerts)} assets requiring immediate attention.",
                "Transfer Status Alert",
                {'Transfer Status Details': immediate_alerts}
            )
        
        # Send scheduled alerts every 15 minutes
        time_since_last_alert = (current_time - monitor.last_status_alert_time).total_seconds() / 60
        if status_alerts and time_since_last_alert >= monitor.config['intervals']['transfer_status_alert_minutes']:
            monitor.send_email(
                "Transfer Status Alert - Scheduled Check",
                f"Transfer status issues detected for {len(status_alerts)} assets during scheduled monitoring.",
                "Transfer Status Report",
                {'Transfer Status Details': status_alerts}
            )
            monitor.last_status_alert_time = current_time
        
        # Check for files older than threshold (keep existing functionality)
        threshold = datetime.now() - timedelta(hours=monitor.config.get('alert_threshold_hours', 1))
        for item in monitor.dropped_files:
            if item['time'] < threshold and os.path.exists(item['path']):
                details = {
                    'Source Folder': item['source_folder'],
                    'Asset Name': item['asset_name'],
                    'Drop Time': item['time'].strftime('%Y-%m-%d %H:%M:%S'),
                    'Duration': str(datetime.now() - item['time'])
                }
                monitor.send_email(
                    "File Stuck Alert",
                    f"File/folder stuck for >1 hour in {item['source_folder']} watch folder.",
                    "Stuck File Alert",
                    details
                )

def report_worker(monitor):
    while monitor.running:
        time.sleep(monitor.config['intervals']['report_minutes'] * 60)
        
        # Generate structured report data
        dropped_summary = []
        for folder_name in monitor.config['watch_folders'].keys():
            folder_drops = [f for f in monitor.dropped_files if f['source_folder'] == folder_name]
            dropped_summary.append({'Source': folder_name, 'Files Count': len(folder_drops)})
        
        dest_counts = []
        total_transfers = 0
        for dest_name, dest_path in monitor.config['destination_folders'].items():
            count = monitor.count_files_in_folder(dest_path)
            dest_counts.append({'Destination': dest_name, 'Files Count': count})
            total_transfers += count
        
        # Create structured details for HTML template
        report_details = {
            'Dropped Files Summary Count at Source': dropped_summary,
            'Transferred Files Summary Count at Destination': dest_counts,
            'Total Transfers': total_transfers
        }
        
        report_body = f"NBA Asset Transfer Monitoring Details - {len(monitor.dropped_files)} total items monitored, {total_transfers} transfers completed."
        
        monitor.send_email("NBA MELTS Transfer Monitoring Notification", report_body, "Monitor Report", report_details)

def performance_report_worker(monitor):
    while monitor.running:
        time.sleep(monitor.config['intervals']['performance_report_minutes'] * 60)
        
        monitor.create_excel_report()
        
        if os.path.exists(monitor.excel_file):
            wb = load_workbook(monitor.excel_file)
            ws = wb.active
            
            report_data = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0]:
                    report_data.append({
                        'Asset Name': row[0],
                        'Asset Type': row[1], 
                        'Count': row[2],
                        'Start Time': row[3],
                        'End Time': row[4],
                        'Total Time': f"{row[5]} min"
                    })
            
            report_body = f"Performance Report - Total {len(report_data)} transfers tracked. See attached Excel file."
            
            monitor.send_email(
                "NBA Performance Report", 
                report_body, 
                "Performance Report", 
                {'Total Transfers': len(report_data)},
                attach_excel=True
            )

def main():
    monitor = FileMonitor()
    
    # Start all worker threads
    validation_thread = threading.Thread(target=validation_worker, args=(monitor,))
    validation_thread.daemon = True
    validation_thread.start()
    
    report_thread = threading.Thread(target=report_worker, args=(monitor,))
    report_thread.daemon = True
    report_thread.start()
    
    performance_thread = threading.Thread(target=performance_report_worker, args=(monitor,))
    performance_thread.daemon = True
    performance_thread.start()
    
    polling_thread = threading.Thread(target=polling_worker, args=(monitor,))
    polling_thread.daemon = True
    polling_thread.start()
    
    print("NBA File Monitor started...")
    
    try:
        while True:
            time.sleep(1)
    except KeyboardInterrupt:
        monitor.running = False

if __name__ == "__main__":
    main()