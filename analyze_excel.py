from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Load the sample Excel file
wb = load_workbook(r"D:\Documents\Automation\SamplePyProgramming\AIMLPrograms\NBA\NBA Auto Generated Performance Report.xlsx")
ws = wb.active

print(f"Sheet name: {ws.title}")
print(f"Max row: {ws.max_row}")
print(f"Max column: {ws.max_column}")

# Check headers and formatting
print("\nHeaders and formatting:")
for col in range(1, ws.max_column + 1):
    cell = ws.cell(row=1, column=col)
    print(f"Column {col}: '{cell.value}' - Bold: {cell.font.bold if cell.font else False}")

# Check data rows
print("\nSample data rows:")
for row in range(2, min(ws.max_row + 1, 5)):  # Show first 3 data rows
    row_data = []
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        row_data.append(cell.value)
    print(f"Row {row}: {row_data}")

# Check column widths
print("\nColumn dimensions:")
for col_letter, dimension in ws.column_dimensions.items():
    if dimension.width:
        print(f"Column {col_letter}: width = {dimension.width}")